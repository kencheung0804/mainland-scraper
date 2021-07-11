from datetime import datetime

from openpyxl import load_workbook
from typing import Optional

from constants.index import source_number_location_dict, line_breaker
from scrapers.extract import find_right_combination
from scrapers.prepare import scrape_city_link_dict, init_new_driver, find_link_el_for_house_type, choose_params
from scrapers.store import save_result_ws, save_target_ws
from utils.common import input_loop
from utils.find_from_excel import find_start_row, find_price_ranges
from datetime import datetime
import logging
import argparse


def main(headless_mode: Optional[bool] = False):
    result_filename = input_loop("The name of result file (e.g. test.xlsx):", "[a-zA-Z0-9]*.xlsx",
                                 "The name should end with .xlsx")
    target_filename = input_loop("The name of target file (e.g. C1.xlsx):", "[a-zA-Z0-9]*.xlsx",
                                 "The name should end with .xlsx")
    search_date_text = input_loop("Your search date (e.g. 2021-01-31):", "[0-9]{4}-[0-9]{2}-[0-9]{2}",
                                  "Date format is not correct")
    search_date = datetime.strptime(search_date_text, '%Y-%m-%d').date()
    wb = load_workbook(target_filename)
    ws = wb['C1p.1']
    start_row = find_start_row(ws)
    city_link_dict = scrape_city_link_dict()

    while ws[f'H{start_row}'].value is not None:
        name_of_property = ws[f'G{start_row}'].value
        address = ws[f'H{start_row}'].value
        city, region = [a.strip() for a in address.split(', ')]
        house_type_list = [t.strip().lower() for t in ws[f'I{start_row}'].value.split('/') if
                           t.strip().lower() in list(source_number_location_dict.keys())]
        driver = init_new_driver(headless_mode)
        driver.implicitly_wait(30)
        driver.get(city_link_dict[city])

        for house_type in house_type_list:
            error_msg = f'Cannot find the data for {name_of_property} ({house_type}) in {city} on row {start_row}'
            try:
                logging.info(
                    f'Start searching for {name_of_property} ({house_type}) in {city} on row {start_row}')
                print(f'Start searching for {name_of_property} ({house_type}) in {city} on row {start_row}')
                driver.implicitly_wait(30)
                max_price, min_price, max_price_for_each, min_price_for_each = find_price_ranges(ws, start_row,
                                                                                                 house_type)

                driver.implicitly_wait(30)

                el = find_link_el_for_house_type(driver, house_type)
                el.click()
                driver.implicitly_wait(30)
                choose_params(driver, house_type, region, min_price_for_each, max_price_for_each)

                driver.implicitly_wait(30)
                price_list, image_list = find_right_combination(driver, house_type, min_price, max_price, search_date,
                                                                min_price_for_each, max_price_for_each)

                if not (price_list is None or image_list is None):
                    save_result_ws(result_filename, price_list, image_list, name_of_property, house_type)
                    save_target_ws(wb, ws, target_filename, house_type, start_row, sum(price_list) / 3 * 0.9)
                    logging.info(
                        f'Done searching for {name_of_property} ({house_type}) in {city} on row {start_row}' + line_breaker)
                    print(
                        f'Done searching for {name_of_property} ({house_type}) in {city} on row {start_row}' + line_breaker)
                else:
                    logging.info(error_msg + line_breaker)
                    print(error_msg + line_breaker)
            except Exception as e:
                print(e)
                print(error_msg + line_breaker)
                logging.error(error_msg + line_breaker)
        driver.close()
        start_row += 1


if __name__ == '__main__':
    logging.basicConfig(filename=f"logbook_{datetime.today().strftime('%m_%d_%Y_%I_%M_%S')}.txt",
                        level=logging.DEBUG,
                        format='%(levelname)s: %(asctime)s %(message)s',
                        datefmt='%m/%d/%Y %I:%M:%S')

    parser = argparse.ArgumentParser()
    parser.add_argument('-H', "--headless_mode", help="Run chrome driver in headless mode",
                        action='store_true')
    args = parser.parse_args()
    main(args.headless_mode)
