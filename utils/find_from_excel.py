import logging

from openpyxl.worksheet.worksheet import Worksheet

from constants.index import source_number_location_dict, FormatError


def find_price_ranges(ws: Worksheet, row: int, house_type: str):
    try:
        raw_valuation_col = source_number_location_dict[house_type]['valuation']
        raw_actual_col = source_number_location_dict[house_type]['actual']
        raw_valuation = str(ws[f'{raw_valuation_col}{row}'].value)
        raw_actual = str(ws[f'{raw_actual_col}{row}'].value)
        max_price_in_valuation = max(sorted([float(v.strip().replace(',', '')) for v in raw_valuation.split('-')]))
        min_price_in_valuation = min(sorted([float(v.strip().replace(',', '')) for v in raw_valuation.split('-')]))
        max_price_in_actual = max(sorted([float(v.strip().replace(',', '')) for v in raw_actual.split('-')]))
        min_price_in_actual = min(sorted([float(v.strip().replace(',', '')) for v in raw_actual.split('-')]))

        all_list = [max_price_in_valuation, min_price_in_valuation, max_price_in_actual, min_price_in_actual]
        min_price_for_each = min(all_list) * 0.85
        max_price_for_each = max(all_list) * 1.15

        v_max = max_price_in_valuation * 1.1
        v_min = min_price_in_valuation * 0.9
        a_max = max_price_in_actual * 1.1
        a_min = max_price_in_actual * 0.9

        if a_max <= v_min or v_max <= a_min:
            upper_limit = v_max
            lower_limit = v_min
        elif a_min >= v_min and a_max <= v_max:
            upper_limit = v_max
            lower_limit = v_min
        elif a_min <= v_min <= a_max <= v_max:
            upper_limit = a_max
            lower_limit = v_min
        elif v_min <= a_min <= v_max <= a_max:
            upper_limit = v_max
            lower_limit = a_min
        elif a_max >= v_max and v_min >= a_min:
            upper_limit = v_max
            lower_limit = v_min

        return upper_limit, lower_limit, max_price_for_each, min_price_for_each
    except Exception:
        error_msg = "Please check your number format in valuation and actual price. And please make sure the number matches with the Usage column."
        logging.error(error_msg)
        print(error_msg)
        raise FormatError(error_msg)


def find_start_row(ws: Worksheet):
    for i in range(1, ws.max_row + 1):
        if ws[f'G{i}'].value == 'Mainland investment properties:':
            return i + 1
    error_msg = "Please check your excel format. The program is searching for the line of 'Mainland investment properties:' to start with."
    logging.error(error_msg)
    print(error_msg)
    raise FormatError(error_msg)
