from io import BytesIO
from openpyxl import Workbook, load_workbook
import openpyxl
import base64
from PIL import Image
from openpyxl.worksheet.worksheet import Worksheet

from constants.index import Number, result_number_location_dict


def save_result_ws(filename: str, price_list: list[Number], image_list: list[str], name_of_property: str,
                   house_type: str,
                   img_height: int = 400, img_width: int = 1000):
    try:
        wb = load_workbook(filename)
    except:
        wb = Workbook()
    ws = wb.active
    ws[f'A{ws.max_row + 5}'] = f"{name_of_property}'s {house_type}:"
    last_row = ws.max_row + 1
    for raw_img in image_list:
        im = Image.open(BytesIO(base64.b64decode(raw_img)))
        img = openpyxl.drawing.image.Image(im)
        img.anchor = f'A{last_row}'
        img.height = img_height
        img.width = img_width
        ws.add_image(img)
        ws.row_dimensions[last_row].height = img.height
        last_row += 1
    ws[f'A{last_row}'] = f'(source: {"fang.com"})'

    ws[f'A{ws.max_row + 2}'] = 'Average price:'
    ws[f'A{ws.max_row + 1}'] = f"({'+'.join([str(p) for p in price_list])})/3 * 0.9"
    ws[f'A{ws.max_row + 1}'] = f'{sum(price_list) * 0.9 / 3}'

    wb.save(filename)


def save_target_ws(wb: openpyxl.workbook.workbook.Workbook, ws: Worksheet, filename: str, house_type: str, row: int,
                   value: float):
    target_column = result_number_location_dict[house_type]
    ws[f'{target_column}{row}'].value = value
    wb.save(filename)
